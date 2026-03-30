"""
Ventana principal de la aplicacion Generador de Planos Forestales.

Layout moderno: 1100x780 px minimo, redimensionable.
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from .estilos import (
    COLOR_FONDO_APP, COLOR_ACENTO, COLOR_ACENTO2, COLOR_TEXTO_GRIS,
    COLOR_PANEL, COLOR_TEXTO, COLOR_BORDE, COLOR_HEADER, COLOR_ENTRY,
    FONT_BOLD, FONT_MONO, FONT_SMALL, FONT_SUBTITULO, FONT_SECCION,
    FONT_BOTON,
    aplicar_estilos, crear_boton,
)
from .panel_capas import PanelCapas
from .panel_config import PanelConfig
from .panel_campos import PanelCampos
from .panel_filtros import PanelFiltros
from .panel_simbologia import PanelSimbologia
from .panel_cajetin import PanelCajetin
from .panel_generacion import PanelGeneracion
from ..motor.generador import GeneradorPlanos
from ..motor.proyecto import Proyecto


class App(tk.Tk):
    """Ventana principal de la aplicacion."""

    def __init__(self):
        super().__init__()
        self.title("Generador de Planos Forestales")
        self.geometry("1160x820")
        self.minsize(1100, 780)
        self.configure(bg=COLOR_FONDO_APP)
        self.resizable(True, True)

        self.motor = GeneradorPlanos()
        self._ultimo_dir_proyecto = os.path.expanduser("~")

        aplicar_estilos(self)
        self._construir_ui()

    def _construir_ui(self):
        # ── Barra superior ──
        self._barra_superior()

        # ── Contenedor principal ──
        main = tk.Frame(self, bg=COLOR_FONDO_APP)
        main.pack(fill="both", expand=True, padx=0, pady=0)

        # Columna izquierda (controles) con scroll
        izq_container = tk.Frame(main, bg=COLOR_PANEL, bd=0,
                                  highlightthickness=0)
        izq_container.pack(side="left", fill="y", padx=(8, 0), pady=8)

        izq_canvas = tk.Canvas(izq_container, bg=COLOR_PANEL,
                                highlightthickness=0, width=300)
        izq_scrollbar = ttk.Scrollbar(izq_container, orient="vertical",
                                       command=izq_canvas.yview)

        izq = tk.Frame(izq_canvas, bg=COLOR_PANEL, bd=0)
        izq.bind("<Configure>",
                 lambda e: izq_canvas.configure(scrollregion=izq_canvas.bbox("all")))
        izq_canvas.create_window((0, 0), window=izq, anchor="nw")
        izq_canvas.configure(yscrollcommand=izq_scrollbar.set)

        # Mousewheel scroll
        def _on_mousewheel(event):
            izq_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_mousewheel_linux(event):
            if event.num == 4:
                izq_canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                izq_canvas.yview_scroll(1, "units")
        izq_canvas.bind("<MouseWheel>", _on_mousewheel)
        izq_canvas.bind("<Button-4>", _on_mousewheel_linux)
        izq_canvas.bind("<Button-5>", _on_mousewheel_linux)

        izq_canvas.pack(side="left", fill="both", expand=True)
        izq_scrollbar.pack(side="right", fill="y")

        # Columna derecha (tabla + log)
        der = tk.Frame(main, bg=COLOR_FONDO_APP)
        der.pack(side="right", fill="both", expand=True, padx=8, pady=8)

        # ── Paneles izquierda (orden de flujo de trabajo) ──

        # 1. Carga de datos
        self.panel_capas = PanelCapas(
            izq, self.motor,
            callback_log=self._escribir_log,
            callback_tabla=self._on_tabla_cargada,
            callback_montes_cargados=self._on_montes_cargados,
        )

        # 2. Filtrado de datos
        self.panel_filtros = PanelFiltros(
            izq, self.motor,
            callback_filtro=self._on_filtro_aplicado,
        )

        # 3. Estilo visual
        self.panel_simbologia = PanelSimbologia(
            izq, self.motor,
            callback_log=self._escribir_log,
        )

        # 4. Campos a mostrar en el plano
        self.panel_campos = PanelCampos(izq)

        # 5. Cajetin y plantilla
        self.panel_cajetin = PanelCajetin(
            izq, self.motor,
            callback_log=self._escribir_log,
        )

        # 6. Configuracion de salida
        self.panel_config = PanelConfig(izq)

        # 7. Generacion final
        self.panel_generacioneracion = PanelGeneracion(
            izq, self.motor,
            get_config=self._get_config,
            callback_log=self._escribir_log,
            auto_aplicar=self._auto_aplicar_todo,
        )

        # ── Panel derecho: tabla + log ──
        self._crear_panel_tabla(der)
        self._crear_panel_log(der)

    def _barra_superior(self):
        barra = tk.Frame(self, bg=COLOR_HEADER, height=56)
        barra.pack(fill="x")
        barra.pack_propagate(False)

        # Contenido izquierdo: titulo
        left = tk.Frame(barra, bg=COLOR_HEADER)
        left.pack(side="left", fill="y")

        tk.Label(
            left, text="\U0001f5fa",
            font=("Segoe UI", 18), bg=COLOR_HEADER, fg=COLOR_ACENTO,
        ).pack(side="left", padx=(16, 6), pady=8)

        title_frame = tk.Frame(left, bg=COLOR_HEADER)
        title_frame.pack(side="left", pady=8)

        tk.Label(
            title_frame, text="Generador de Planos Forestales",
            font=("Segoe UI", 13, "bold"), bg=COLOR_HEADER, fg=COLOR_TEXTO,
        ).pack(anchor="w")

        tk.Label(
            title_frame,
            text="\u00a9 Jose Caballero S\u00e1nchez \u00b7 Cazorla 2026",
            font=("Segoe UI", 8), bg=COLOR_HEADER, fg=COLOR_TEXTO_GRIS,
        ).pack(anchor="w")

        # Contenido derecho: botones de proyecto + CRS
        right = tk.Frame(barra, bg=COLOR_HEADER)
        right.pack(side="right", fill="y", padx=12)

        # Badge CRS
        crs_frame = tk.Frame(right, bg=COLOR_BORDE, bd=0,
                              highlightthickness=0)
        crs_frame.pack(side="right", pady=14, padx=(8, 0))
        tk.Label(
            crs_frame, text=" ETRS89 \u00b7 UTM H30N ",
            font=("Segoe UI", 8, "bold"), bg=COLOR_BORDE, fg=COLOR_TEXTO_GRIS,
        ).pack(padx=6, pady=2)

        # Botones de proyecto
        btn_f = tk.Frame(right, bg=COLOR_HEADER)
        btn_f.pack(side="right", pady=10)

        btn_guardar = tk.Button(
            btn_f, text="\U0001f4be  Guardar", command=self._guardar_proyecto,
            font=FONT_BOTON, bg=COLOR_BORDE, fg=COLOR_TEXTO,
            relief="flat", cursor="hand2", padx=10, pady=3,
            bd=0, highlightthickness=0,
            activebackground=COLOR_ACENTO, activeforeground="#FFFFFF")
        btn_guardar.pack(side="left", padx=(0, 6))

        btn_cargar = tk.Button(
            btn_f, text="\U0001f4c2  Cargar", command=self._cargar_proyecto,
            font=FONT_BOTON, bg=COLOR_BORDE, fg=COLOR_TEXTO,
            relief="flat", cursor="hand2", padx=10, pady=3,
            bd=0, highlightthickness=0,
            activebackground=COLOR_ACENTO, activeforeground="#FFFFFF")
        btn_cargar.pack(side="left")

        # Hover effects
        for btn in (btn_guardar, btn_cargar):
            btn.bind("<Enter>", lambda e, b=btn: b.configure(bg="#2A4058"))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(bg=COLOR_BORDE))

    def _crear_panel_tabla(self, parent):
        """Crea la tabla de infraestructuras."""
        # Header de seccion
        header = tk.Frame(parent, bg=COLOR_FONDO_APP)
        header.pack(fill="x", pady=(0, 4))

        tk.Label(
            header, text="\U0001f4ca  INFRAESTRUCTURAS",
            font=FONT_SECCION, bg=COLOR_FONDO_APP, fg=COLOR_ACENTO,
        ).pack(side="left")

        # Contenedor de tabla
        tabla_container = tk.Frame(parent, bg=COLOR_PANEL, bd=0,
                                    highlightthickness=1,
                                    highlightbackground=COLOR_BORDE)
        tabla_container.pack(fill="both", expand=True, pady=(0, 6))

        cols = ["#"]
        self._tabla = ttk.Treeview(tabla_container, columns=cols,
                                    show="headings", selectmode="extended")
        self._tabla.heading("#", text="#")
        self._tabla.column("#", width=60, minwidth=40)

        sb_v = ttk.Scrollbar(tabla_container, orient="vertical",
                              command=self._tabla.yview)
        sb_h = ttk.Scrollbar(tabla_container, orient="horizontal",
                              command=self._tabla.xview)
        self._tabla.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)

        sb_h.pack(side="bottom", fill="x")
        sb_v.pack(side="right", fill="y")
        self._tabla.pack(side="left", fill="both", expand=True)

    def _reconfigurar_tabla(self, columnas: list):
        """Reconfigura las columnas de la tabla."""
        cols = ["#"] + columnas
        self._tabla.configure(columns=cols)
        for col in cols:
            ancho = 50 if col == "#" else 120
            self._tabla.heading(col, text=col)
            self._tabla.column(col, width=ancho, minwidth=40)

    def _crear_panel_log(self, parent):
        # Header de seccion
        header = tk.Frame(parent, bg=COLOR_FONDO_APP)
        header.pack(fill="x", pady=(4, 4))

        tk.Label(
            header, text="\U0001f4dd  LOG DE PROCESO",
            font=FONT_SECCION, bg=COLOR_FONDO_APP, fg=COLOR_ACENTO,
        ).pack(side="left")

        # Contenedor de log
        log_container = tk.Frame(parent, bg=COLOR_ENTRY, bd=0,
                                  highlightthickness=1,
                                  highlightbackground=COLOR_BORDE)
        log_container.pack(fill="x", pady=(0, 0))

        self._log = tk.Text(
            log_container, height=7, font=FONT_MONO,
            bg=COLOR_ENTRY, fg="#6EE7B7",
            insertbackground=COLOR_ACENTO, relief="flat", state="disabled",
            padx=10, pady=8, bd=0,
            selectbackground=COLOR_ACENTO, selectforeground="#FFFFFF",
        )
        sb = ttk.Scrollbar(log_container, orient="vertical",
                            command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        self._log.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Tags de color
        self._log.tag_config("error", foreground="#FCA5A5")
        self._log.tag_config("ok", foreground="#6EE7B7")
        self._log.tag_config("warn", foreground="#FCD34D")
        self._log.tag_config("info", foreground="#93C5FD")

        self._escribir_log("Sistema iniciado. Carga un shapefile para comenzar.", "info")

    def _escribir_log(self, msg: str, tipo: str = ""):
        def _do():
            self._log.configure(state="normal")
            tag = tipo if tipo else ""
            self._log.insert("end", msg + "\n", tag)
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _do)

    def _poblar_tabla(self, indices=None):
        gdf = self.motor.gdf_infra
        if gdf is None:
            return

        children = self._tabla.get_children()
        if children:
            self._tabla.delete(*children)

        self._tabla.tag_configure("par", background="#162230")
        self._tabla.tag_configure("impar", background="#0F1923")

        columnas = [c for c in gdf.columns if c != "geometry"]

        if indices is None:
            indices = range(len(gdf))

        for i in indices:
            row = gdf.iloc[i]
            vals = [i + 1]
            for col in columnas:
                v = row.get(col)
                vals.append("\u2014" if v is None or str(v) == "nan" else str(v))
            tag = "par" if i % 2 == 0 else "impar"
            self._tabla.insert("", "end", values=vals, tags=(tag,))

        n = len(list(indices)) if not isinstance(indices, range) else len(indices)
        self._escribir_log(f"Tabla actualizada: {n} infraestructuras.", "info")

    def _on_tabla_cargada(self):
        columnas = self.motor.obtener_columnas_shapefile()
        self._reconfigurar_tabla(columnas)
        self._poblar_tabla()
        self.panel_generacioneracion.actualizar_campos_agrupacion()
        self.panel_generacioneracion.actualizar_valores_si_agrupado()
        self.panel_filtros.actualizar_campos()
        self.panel_simbologia.actualizar_capas_extra()
        self.panel_simbologia.actualizar_campo_categoria()
        self.panel_simbologia.actualizar_campo_categoria_montes()
        self.panel_campos.actualizar_campos(columnas)
        self.panel_cajetin.actualizar_campos_subtitulo(columnas)
        self.panel_config.actualizar_campos_shp_enlace(columnas)

        if hasattr(self, "_campos_visibles_proyecto") and self._campos_visibles_proyecto:
            campos_proy = self._campos_visibles_proyecto
            for campo, var in self.panel_campos._check_campos.items():
                var.set(campo in campos_proy)
            self.panel_campos._actualizar_count()
            self._campos_visibles_proyecto = []

    def _on_montes_cargados(self):
        self.panel_simbologia.actualizar_campo_categoria_montes()
        if self.motor.gdf_montes is not None:
            cols = [c for c in self.motor.gdf_montes.columns
                    if c.lower() != "geometry"]
            self.panel_cajetin.actualizar_campos_montes(cols)

    def _on_filtro_aplicado(self, indices: list):
        self._poblar_tabla(indices)

    def _auto_aplicar_todo(self):
        """Aplica cajetin, plantilla, layout y simbologia al motor antes de generar."""
        cajetin = self.panel_cajetin.obtener_cajetin()
        plantilla = self.panel_cajetin.obtener_plantilla()
        self.motor.set_cajetin(cajetin)
        self.motor.set_plantilla(plantilla)
        self.motor.layout_key = self.panel_cajetin.obtener_layout_key()
        self.motor.dpi_figura = self.panel_config.dpi_figura
        self.motor.dpi_guardado = self.panel_config.dpi_guardado
        self.panel_simbologia._aplicar()
        self.motor.config_infra["alpha"] = self.panel_capas.transparencia_infra.get()
        self.motor.ruta_raster_general = self.panel_config.ruta_raster_general
        self.motor.ruta_raster_localizacion = self.panel_config.ruta_raster_localizacion
        self.motor.ruta_capa_localizacion = self.panel_config.ruta_capa_localizacion
        self.motor.wms_custom_general = self.panel_config.wms_custom_general
        self.motor.wfs_custom_general = self.panel_config.wfs_custom_general
        self.motor.wms_custom_localizacion = self.panel_config.wms_custom_localizacion
        self.motor.wfs_custom_localizacion = self.panel_config.wfs_custom_localizacion
        self.motor.escala_localizacion = self.panel_config.escala_localizacion
        self.motor.prov_localizacion = self.panel_config._prov_localizacion.get()
        if self.panel_config.usa_excel and self.panel_config.ruta_excel:
            try:
                self.motor.cargar_excel_tabla(
                    self.panel_config.ruta_excel,
                    hoja=self.panel_config.hoja_excel or None,
                    campo_enlace_shp=self.panel_config.campo_enlace_shp,
                    campo_enlace_excel=self.panel_config.campo_enlace_excel,
                    columnas_activas=self.panel_config.columnas_excel_activas)
            except Exception as e:
                self._escribir_log(f"Error al cargar Excel: {e}", "error")
                self.motor.limpiar_excel_tabla()
        else:
            self.motor.limpiar_excel_tabla()

    def _get_config(self) -> dict:
        return {
            "formato": self.panel_config.formato.get(),
            "proveedor": self.panel_config.proveedor.get(),
            "ruta_raster_general": self.panel_config.ruta_raster_general,
            "ruta_raster_localizacion": self.panel_config.ruta_raster_localizacion,
            "transparencia": self.panel_capas.transparencia.get(),
            "campos": self.panel_campos.obtener_campos_activos(),
            "campo_encabezado": self.panel_campos.obtener_campo_encabezado(),
            "color_infra": self.panel_config.color_infra,
            "salida": self.panel_config.salida.get(),
            "patron_nombre": self.panel_config.patron_nombre.get(),
            "escala_manual": self.panel_config.escala_manual,
            "tabla": self._tabla,
        }

    # ── Guardar/Cargar proyecto ──────────────────────────────────────────

    def _guardar_proyecto(self):
        ruta = filedialog.asksaveasfilename(
            title="Guardar proyecto",
            defaultextension=".json",
            filetypes=[("Proyecto JSON", "*.json")],
            initialdir=self._ultimo_dir_proyecto,
        )
        if not ruta:
            return
        self._ultimo_dir_proyecto = os.path.dirname(ruta)

        try:
            p = Proyecto()
            p.nombre = os.path.splitext(os.path.basename(ruta))[0]
            p.formato = self.panel_config.formato.get()
            p.proveedor = self.panel_config.proveedor.get()
            p.ruta_raster_general = self.panel_config.ruta_raster_general
            p.ruta_raster_localizacion = self.panel_config.ruta_raster_localizacion
            p.prov_localizacion = self.panel_config._prov_localizacion.get()
            p.escala_localizacion = self.panel_config.escala_localizacion
            p.ruta_capa_localizacion = self.panel_config.ruta_capa_localizacion
            p.wms_custom_general = self.panel_config.wms_custom_general
            p.wfs_custom_general = self.panel_config.wfs_custom_general
            p.wms_custom_localizacion = self.panel_config.wms_custom_localizacion
            p.wfs_custom_localizacion = self.panel_config.wfs_custom_localizacion
            p.escala_manual = self.panel_config.escala_manual
            p.transparencia_montes = self.panel_capas.transparencia.get()
            p.transparencia_infra = self.panel_capas.transparencia_infra.get()
            p.color_infra = self.panel_config.color_infra
            p.calidad_pdf = self.panel_config.calidad_pdf
            p.campos_visibles = self.panel_campos.obtener_campos_activos()
            p.campo_encabezado = self.panel_campos.obtener_campo_encabezado() or ""
            p.carpeta_salida = self.panel_config.salida.get()
            p.patron_nombre = self.panel_config.patron_nombre.get()
            p.layout_key = self.panel_cajetin.obtener_layout_key()
            p.cajetin = self.panel_cajetin.obtener_cajetin()
            p.plantilla = self.panel_cajetin.obtener_plantilla()
            p.simbologia = self.motor.gestor_simbologia.to_dict()
            p.capas_extra = self.motor.gestor_capas.to_dict()
            p.origen_datos_tabla = self.panel_config._origen_datos.get()
            p.ruta_excel_tabla = self.panel_config.ruta_excel
            p.hoja_excel_tabla = self.panel_config.hoja_excel
            p.campo_enlace_shp = self.panel_config.campo_enlace_shp
            p.campo_enlace_excel = self.panel_config.campo_enlace_excel
            p.columnas_excel_activas = self.panel_config.columnas_excel_activas

            p.modo_gen = self.panel_generacion._modo_gen.get()
            try:
                p.rango_desde = int(self.panel_generacion._rango_desde.get())
            except ValueError:
                p.rango_desde = 1
            try:
                p.rango_hasta = int(self.panel_generacion._rango_hasta.get())
            except ValueError:
                p.rango_hasta = 10
            p.campo_agrupacion = self.panel_generacion._campo_agrupacion.get()
            p.multipagina = self.panel_generacion._multipagina.get()
            p.incluir_portada = self.panel_generacion._incluir_portada.get()

            p.guardar(ruta)
            self._escribir_log(f"Proyecto guardado: {ruta}", "ok")
        except Exception as e:
            self._escribir_log(f"Error al guardar proyecto: {e}", "error")
            messagebox.showerror("Error", str(e))

    def _cargar_proyecto(self):
        ruta = filedialog.askopenfilename(
            title="Cargar proyecto",
            filetypes=[("Proyecto JSON", "*.json")],
            initialdir=self._ultimo_dir_proyecto,
        )
        if not ruta:
            return
        self._ultimo_dir_proyecto = os.path.dirname(ruta)

        try:
            p = Proyecto.cargar(ruta)

            # ── Configuracion general ──
            self.panel_config.formato.set(p.formato)
            self.panel_config.proveedor.set(p.proveedor)
            if p.ruta_raster_general:
                self.panel_config._ruta_raster.set(p.ruta_raster_general)
                self.panel_config._lbl_raster.configure(
                    text=os.path.basename(p.ruta_raster_general))
            if p.ruta_raster_localizacion:
                self.panel_config._ruta_raster_loc.set(p.ruta_raster_localizacion)
                self.panel_config._lbl_raster_loc.configure(
                    text=os.path.basename(p.ruta_raster_localizacion))
            if hasattr(p, "prov_localizacion") and p.prov_localizacion:
                self.panel_config._prov_localizacion.set(p.prov_localizacion)
            if hasattr(p, "escala_localizacion") and p.escala_localizacion:
                self.panel_config._escala_localizacion.set(
                    f"{p.escala_localizacion:,}")
            if hasattr(p, "ruta_capa_localizacion") and p.ruta_capa_localizacion:
                self.panel_config._ruta_capa_loc.set(p.ruta_capa_localizacion)
                self.panel_config._lbl_capa_loc.configure(
                    text=os.path.basename(p.ruta_capa_localizacion))
            if hasattr(p, "wms_custom_general") and p.wms_custom_general:
                self.panel_config._wms_url.set(p.wms_custom_general.get("url", ""))
                self.panel_config._wms_capa.set(p.wms_custom_general.get("capa", ""))
                self.panel_config._wms_formato.set(
                    p.wms_custom_general.get("formato", "image/png"))
            if hasattr(p, "wfs_custom_general") and p.wfs_custom_general:
                self.panel_config._wfs_url.set(p.wfs_custom_general.get("url", ""))
                self.panel_config._wfs_capa.set(p.wfs_custom_general.get("capa", ""))
            if hasattr(p, "wms_custom_localizacion") and p.wms_custom_localizacion:
                self.panel_config._wms_loc_url.set(
                    p.wms_custom_localizacion.get("url", ""))
                self.panel_config._wms_loc_capa.set(
                    p.wms_custom_localizacion.get("capa", ""))
                self.panel_config._wms_loc_formato.set(
                    p.wms_custom_localizacion.get("formato", "image/png"))
            if hasattr(p, "wfs_custom_localizacion") and p.wfs_custom_localizacion:
                self.panel_config._wfs_loc_url.set(
                    p.wfs_custom_localizacion.get("url", ""))
                self.panel_config._wfs_loc_capa.set(
                    p.wfs_custom_localizacion.get("capa", ""))
            self.panel_config._on_proveedor_changed()
            self.panel_config._on_prov_loc_changed()
            self.panel_config.salida.set(p.carpeta_salida)
            if p.patron_nombre:
                self.panel_config.patron_nombre.set(p.patron_nombre)

            if p.escala_manual:
                self.panel_config._escala_manual.set(f"{p.escala_manual:,}")
            else:
                self.panel_config._escala_manual.set("0 (auto)")

            if p.color_infra:
                self.panel_config._color_infra = p.color_infra
                self.panel_config._lbl_color.configure(bg=p.color_infra)

            if hasattr(p, "calidad_pdf") and p.calidad_pdf:
                self.panel_config._calidad_pdf.set(p.calidad_pdf)

            # ── Transparencias ──
            self.panel_capas.transparencia.set(p.transparencia_montes)
            if hasattr(p, "transparencia_infra"):
                self.panel_capas.transparencia_infra.set(p.transparencia_infra)

            # ── Campos ──
            if hasattr(p, "campo_encabezado") and p.campo_encabezado:
                self.panel_campos._combo_encabezado.set(p.campo_encabezado)

            # ── Layout y cajetin ──
            if p.layout_key:
                self.panel_cajetin._layout_key.set(p.layout_key)
                self.motor.layout_key = p.layout_key
            self.panel_cajetin.cargar_desde_proyecto(p.cajetin, p.plantilla)

            self.motor.set_cajetin(p.cajetin)
            self.motor.set_plantilla(p.plantilla)

            # ── Origen datos tabla ──
            if hasattr(p, "origen_datos_tabla") and p.origen_datos_tabla:
                self.panel_config._origen_datos.set(p.origen_datos_tabla)
                self.panel_config._on_origen_datos_changed()
            if hasattr(p, "ruta_excel_tabla") and p.ruta_excel_tabla:
                self.panel_config._ruta_excel.set(p.ruta_excel_tabla)
                self.panel_config._lbl_excel.configure(
                    text=os.path.basename(p.ruta_excel_tabla))
                hoja = getattr(p, "hoja_excel_tabla", "") or ""
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(
                        p.ruta_excel_tabla, read_only=True, data_only=True)
                    self.panel_config._cb_hoja.configure(values=wb.sheetnames)
                    wb.close()
                    if hoja:
                        self.panel_config._hoja_excel.set(hoja)
                    self.panel_config._cargar_columnas_excel(
                        p.ruta_excel_tabla, hoja or wb.sheetnames[0])
                except Exception:
                    pass
            if hasattr(p, "hoja_excel_tabla") and p.hoja_excel_tabla:
                self.panel_config._hoja_excel.set(p.hoja_excel_tabla)
            if hasattr(p, "campo_enlace_shp") and p.campo_enlace_shp:
                self.panel_config._campo_enlace_shp.set(p.campo_enlace_shp)
            if hasattr(p, "campo_enlace_excel") and p.campo_enlace_excel:
                self.panel_config._campo_enlace_excel.set(p.campo_enlace_excel)
            if hasattr(p, "columnas_excel_activas") and p.columnas_excel_activas:
                cols_proy = p.columnas_excel_activas
                for col, var in self.panel_config._check_cols_excel.items():
                    var.set(col in cols_proy)

            # ── Simbologia ──
            if p.simbologia:
                from ..motor.simbologia import GestorSimbologia
                self.motor.gestor_simbologia = GestorSimbologia.from_dict(p.simbologia)

            # ── Generacion ──
            if hasattr(p, "modo_gen") and p.modo_gen:
                self.panel_generacion._modo_gen.set(p.modo_gen)
            if hasattr(p, "rango_desde"):
                self.panel_generacion._rango_desde.delete(0, "end")
                self.panel_generacion._rango_desde.insert(0, str(p.rango_desde))
            if hasattr(p, "rango_hasta"):
                self.panel_generacion._rango_hasta.delete(0, "end")
                self.panel_generacion._rango_hasta.insert(0, str(p.rango_hasta))
            if hasattr(p, "campo_agrupacion") and p.campo_agrupacion:
                self.panel_generacion._campo_agrupacion.set(p.campo_agrupacion)
            if hasattr(p, "multipagina"):
                self.panel_generacion._multipagina.set(p.multipagina)
            if hasattr(p, "incluir_portada"):
                self.panel_generacion._incluir_portada.set(p.incluir_portada)

            self._campos_visibles_proyecto = p.campos_visibles or []

            self._escribir_log(f"Proyecto cargado: {p.nombre}", "ok")
        except Exception as e:
            self._escribir_log(f"Error al cargar proyecto: {e}", "error")
            messagebox.showerror("Error", str(e))
